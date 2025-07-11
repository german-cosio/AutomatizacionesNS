import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import xmlrpc.client
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import Workbook
from utils.emailer import send_email

# Load environment variables
load_dotenv()

# Configuration
CONFIG = {
    'url': os.getenv('url'),
    'db': os.getenv('db'),
    'username': os.getenv('odoo_username'),
    'password': os.getenv('password'),
    'recipients': os.getenv('recipients'),
}

class InvoiceManager:
    def __init__(self, url:str, db:str, username:str, password:str, recipients:str):
        self.url = url
        self.db = db
        self.username = username
        self.password = password
        self.recipients = recipients.split(',')
        self.uid = None
        self.models = None
        print(f"Procesando: 10%")

    def connect(self):
        common = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/common')
        self.uid = common.authenticate(self.db, self.username, self.password, {})
        self.models = xmlrpc.client.ServerProxy(f'{self.url}/xmlrpc/2/object')
        print(f"Procesando: 30%")

    def check_invoice_due_dates(self):
        today = datetime.now().date()
        near_future = today + timedelta(days=7)

        expired_invoices_clients = []
        due_soon_invoices_clients = []
        expired_invoices_prov = []
        due_soon_invoices_prov = []

        expired_invoices_client_count, due_soon_invoices_client_count = 0, 0
        expired_invoices_prov_count, due_soon_invoices_prov_count = 0, 0

        file_path = 'facturas_vencidas_y_por_vencer.xlsx'
        print(f"Procesando: 20%")

        try:
            self.connect()
            
            # Retrieve specified fields from the filtered account moves
            fields = ['invoice_date', 'invoice_date_due', 'amount_total', 'amount_residual', 'invoice_origin', 'invoice_partner_display_name', 'display_name','invoice_payment_term_id', 'state']

            # Search for account moves with a sequence prefix of 'INV/2024/' and payment_state 'not_paid'
            move_ids_inv = self.models.execute_kw(self.db, self.uid, self.password, 
                                                  'account.move', 'search_read', 
                                                  [[['sequence_prefix', '=', 'INV/2024/'], ['payment_state', '=', 'not_paid'], ['state', '=', 'posted']]], 
                                                  {'fields': fields}
                                                  )

            move_ids_factu = self.models.execute_kw(self.db, self.uid, self.password, 
                                                    'account.move', 'search_read', 
                                                    [[['payment_state', '=', 'not_paid'], ['name', 'ilike', 'FACTU/2024/'], ['state', '=', 'posted']]],
                                                    {'fields': fields}
                                                    )
            print(f"Procesando: 40%")

            if move_ids_inv:
                for move in move_ids_inv:
                    due_date = datetime.strptime(move['invoice_date_due'], '%Y-%m-%d').date()
                    if due_date <= today:
                        expired_invoices_clients.append(move)
                        expired_invoices_client_count += 1
                    elif today < due_date <= near_future:
                        due_soon_invoices_clients.append(move)
                        due_soon_invoices_client_count += 1
            print(f"Procesando: 50%")

            if move_ids_factu:
                for move in move_ids_factu:
                    due_date = datetime.strptime(move['invoice_date_due'], '%Y-%m-%d').date()
                    if due_date <= today:
                        expired_invoices_prov.append(move)
                        expired_invoices_prov_count += 1
                    elif today < due_date <= near_future:
                        due_soon_invoices_prov.append(move)
                        due_soon_invoices_prov_count += 1
            print(f"Procesando: 60%")

            # Generar archivo Excel
            create_excel_file(file_path, expired_invoices_clients, due_soon_invoices_clients, expired_invoices_prov, due_soon_invoices_prov)
            print(f"Procesando: 90%")

            # Enviar correos electrónicos
            subject = 'Notificación de Facturas Vencidas y Próximas a Vencer'
            body = (
                f'Facturas vencidas de: Clientes - {expired_invoices_client_count} | Proveedores - {expired_invoices_prov_count}\n'
                f'Facturas por vencer de: Clientes - {due_soon_invoices_client_count} | Proveedores - {due_soon_invoices_prov_count}\n\n'
                f'Adjunto encontrará los reportes de facturas vencidas y próximas a vencer.'
            )
            print(f"Procesando: 100%")

            attachments = [file_path]
            send_email(self.recipients, subject, body, attachments)

        except Exception as e:
            print(f"An error occurred: {e}")

def create_excel_file(file_name, expired_invoices_clients, due_soon_invoices_clients, expired_invoices_prov, due_soon_invoices_prov):
        workbook = Workbook()
        data = [
            ('Vencidas clientes', expired_invoices_clients),
            ('Por vencer Clientes', due_soon_invoices_clients),
            ('Vencidas proveedores', expired_invoices_prov),
            ('Por vencer proveedores', due_soon_invoices_prov)
        ]

        print(f"Procesando: 70%")
        for title, invoices in data:
            sheet = workbook.create_sheet(title=title)
            headers = ['Fecha de factura', 'Fecha de vencimiento', 'Importe total', 'Importe residual', 'Origen', 'Partner Name', 'Display Name', 'Término de pago']
            sheet.append(headers)
            for record in invoices:
                row = [
                    record['invoice_date'],
                    record['invoice_date_due'],
                    record['amount_total'],
                    record['amount_residual'],
                    record['invoice_origin'],
                    record['invoice_partner_display_name'],
                    record['display_name'],
                    record['invoice_payment_term_id'][1] if record['invoice_payment_term_id'] else None
                ]
                sheet.append(row)

            # Ajustar el ancho de las columnas
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[column].width = max_length + 2
        print(f"Procesando: 80%")

        # Remove the default sheet created by openpyxl
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])

        workbook.save(file_name)

# Initialize the InvoiceManager
invoice_manager = InvoiceManager(CONFIG['url'], CONFIG['db'], CONFIG['username'], CONFIG['password'], CONFIG['recipients'])

# Retrieve and print account moves
invoice_manager.check_invoice_due_dates()