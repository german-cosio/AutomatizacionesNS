import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

def extract_column_values(file_path, column):
    app = xw.App(visible=False)
    workbook = app.books.open(file_path)
    sheet = workbook.sheets[0]
    values = sheet.range(f'{column}2:{column}{sheet.cells.last_cell.row}').value
    values = [value for value in values if value is not None]
    workbook.close()
    app.quit()
    return values

def write_to_new_excel(inventario, inventario_nombres, lista_de_precios, output_file):
    cant_faltantes = 0

    workbook = Workbook()
    sheet = workbook.active
    
    red_fill = PatternFill(start_color="FF5232", end_color="FF5232", fill_type="solid")
    
    row = 1
    for ref_int, nombre in zip(inventario, inventario_nombres):
        # Verificar si la referencia interna es de "caja abierta" 
        if "caja abierta" in nombre.lower():
            # Si lo es eliminar "-OB" o "OB-" de la referencia interna
            ref_int = re.sub(r"-OB|OB-", "", ref_int)

        cell = sheet.cell(row=row, column=1, value=ref_int)
        cell_b = sheet.cell(row=row, column=2, value=nombre)
        if ref_int not in lista_de_precios:
            cell.fill = red_fill
            cell_b.fill = red_fill
            cant_faltantes += 1
        row += 1
    
    print(f"{cant_faltantes} productos de Inventario faltantes en Lista de precios")

    workbook.save(output_file)