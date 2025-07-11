import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

def write_invoices_to_excel(file_path, invoices):
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet("ANNUAL_SALES_INVOICES")

    # Add a header format
    header_format = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    headers = ['Invoice Name', 'Invoice Date', 'Total Amount', 'Line_ids', 'Sequence_prefix', 'state', 'move_type', 'currency', 'user', 'payment state', 'invoice origin', 'invoice partner display name']

    for i, header in enumerate(headers):
        worksheet.write(0, i, header, header_format)

    # Writing invoices to Excel
    for row_num, invoice in enumerate(invoices, 1):
        worksheet.write(row_num, 0, invoice['display_name'])
        worksheet.write(row_num, 1, invoice['invoice_date'])
        worksheet.write(row_num, 2, invoice['amount_total'])
        worksheet.write(row_num, 3, str(invoice['line_ids']))
        worksheet.write(row_num, 4, invoice['sequence_prefix'])
        worksheet.write(row_num, 5, invoice['state'])
        worksheet.write(row_num, 6, invoice['move_type'])

        currency_id = invoice.get('currency_id')
        if currency_id and isinstance(currency_id, (list, tuple)) and len(currency_id) > 1:
            worksheet.write(row_num, 7, currency_id[1])
        else:
            worksheet.write(row_num, 7, '')

        user_id = invoice.get('user_id')
        if user_id and isinstance(user_id, (list, tuple)) and len(user_id) > 1:
            worksheet.write(row_num, 8, user_id[1])
        else:
            worksheet.write(row_num, 8, '')

        worksheet.write(row_num, 9, invoice['payment_state'])
        worksheet.write(row_num, 10, invoice['invoice_origin'])
        worksheet.write(row_num, 11, invoice['invoice_partner_display_name'])

    workbook.close()
    print(f"\033[92mANNUAL_SALES_INVOICES añadido con exito\033[0m")

def write_products_to_excel(file_path, data):
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.create_sheet(title="ANNUAL_SALES_PRODUCTS")

    # Add a header format
    headers = ['Product Name', 'Quantity', 'Unit Price', 'Invoice Date', 'Category', 'Invoice Name']
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Writing data to Excel
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(file_path)
    print(f"\033[92mANNUAL_SALES_PRODUCTS añadido con exito\033[0m")