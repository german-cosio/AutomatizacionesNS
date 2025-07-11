import os
import time
import xmlrpc.client
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def get_peps():
    # Define la ruta de la carpeta
    files = os.listdir("stock_valuer\\src\\conteo_de_inventario")
    file_path = os.path.join("stock_valuer\\src\\conteo_de_inventario", files[0])

    df_existing = pd.read_excel(file_path)
    product_quantities = dict(zip(df_existing['Referencia interna'], df_existing['C. inventariada']))

    start_time = time.time()
    print("Starting process to fetch latest invoice data...")

    # Get latest invoice data
    latest_invoices = get_latest_invoice_data(url, db, username, password, product_quantities, fallback_csv)

    # Prepare data for writing to Excel
    results = []
    for product_id, invoices in latest_invoices.items():
        if invoices:
            for invoice in invoices:
                results.append([product_id, invoice['move_id'][0], invoice['move_id'][1], invoice['move_id'][2], invoice['name'], invoice['date'], invoice['quantity'], invoice['price_unit'], invoice['currency_id'][1]])
        else:
            results.append([product_id, 'No invoice found to satisfy the required quantity', 'N/A', 'N/A', 'N/A', product_quantities[product_id], 'N/A', 'N/A'])

    df_results = pd.DataFrame(results, columns=['Product ID', 'Invoice ID', 'Invoice', 'Invoice Total', 'Name', 'Date', 'Quantity', 'Price unit', 'Currency'])

    # Load the existing Excel file
    book = load_workbook(file_path)

    # Create a new sheet for results if it doesn't exist
    if 'RESULTS' in book.sheetnames:
        del book['RESULTS']  # Remove the old RESULTS sheet if it exists
    results_sheet = book.create_sheet('RESULTS')

    # Highlight the cells if the quantity is not met
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Write the DataFrame to the RESULTS sheet, including headers
    for c_idx, col in enumerate(df_results.columns, 1):
        results_sheet.cell(row=1, column=c_idx, value=col)

    for r_idx, row in enumerate(df_results.values, 2):
        for c_idx, value in enumerate(row, 1):
            cell = results_sheet.cell(row=r_idx, column=c_idx, value=value)
            if c_idx == 6 and value == 'N/A':
                cell.fill = highlight

    # Save the workbook
    book.save(file_path)

    end_time = time.time()
    print(f"Process completed in {end_time - start_time:.2f} seconds.")

def get_latest_invoice_data(url, db, username, password, product_quantities, fallback_csv):
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    uid = common.authenticate(db, username, password, {})
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

    # Read the fallback CSV file
    fallback_df = pd.read_csv(fallback_csv)
    
    result = {}
    cont = 0
    cant = len(product_quantities)
    for product_id, required_quantity in product_quantities.items():
        cont += 1
        print(f'Processing {cont}/{cant}', end='\r')
        # Skip products with C. disp = 0
        if required_quantity == 0:
            continue
        
        time.sleep(0.2)
        try:
            data = models.execute_kw(db, uid, password, 'account.move.line', 'search_read',
                                     [[['product_id.default_code', '=', product_id], ['move_name', 'ilike', 'FACTU'], ['date', '<=', '2024-06-03']]],
                                     {'fields': ['id', 'name', 'move_id', 'date', 'quantity', 'price_unit', 'currency_id'], 'order': 'date asc'})

            for line in data:
                time.sleep(0.2)
                invoice_data = models.execute_kw(db, uid, password, 'account.move', 'search_read',
                                     [[['id', '=', line['move_id'][0]]]],
                                     {'fields': ['id', 'name', 'date', 'partner_id', 'line_ids', 'amount_total', 'state']})

                line['move_id'].append(invoice_data[0]['amount_total'])

        except OverflowError as e:
            print(f"\nOverflowError for product ID: {product_id} - {e}")
            continue

        if data:
            total_quantity = 0
            selected_invoices = []
            remaining_quantity = required_quantity

            for record in data:
                if remaining_quantity <= 0:
                    break
                available_quantity = record['quantity']
                if available_quantity >= remaining_quantity:
                    record['quantity'] = remaining_quantity
                    selected_invoices.append(record)
                    remaining_quantity = 0
                else:
                    remaining_quantity -= available_quantity
                    selected_invoices.append(record)
            
            if remaining_quantity <= 0:
                result[product_id] = selected_invoices
            else:
                # Check the fallback CSV for the product_id
                fallback_row = fallback_df[fallback_df['product_id'] == product_id]
                if not fallback_row.empty:
                    result[product_id] = [{
                        'move_id': ['N/A', 'N/A', 'N/A'],
                        'name': 'N/A',
                        'date': 'N/A',
                        'quantity': remaining_quantity,
                        'price_unit': fallback_row['price_unit'].values[0] if 'price_unit' in fallback_row else 'N/A',
                        'currency_id': ['N/A', 'N/A']
                    }]
                else:
                    print(f"\nPrice not set for product ID: {product_id}")
                    result[product_id] = [{
                        'move_id': ['N/A', 'N/A', 'N/A'],
                        'name': 'N/A',
                        'date': 'N/A',
                        'quantity': required_quantity,
                        'price_unit': 'N/A',
                        'currency_id': ['N/A', 'N/A']
                    }]
        else:
            # Check the fallback CSV for the product_id
            fallback_row = fallback_df[fallback_df['product_id'] == product_id]
            if not fallback_row.empty:
                result[product_id] = [{
                    'move_id': ['N/A', 'N/A', 'N/A'],
                    'name': 'N/A',
                    'date': 'N/A',
                    'quantity': required_quantity,
                    'price_unit': fallback_row['price_unit'].values[0] if 'price_unit' in fallback_row else 'N/A',
                    'currency_id': ['N/A', 'N/A']
                }]
            else:
                print(f"\nPrice not set for product ID: {product_id}")
                result[product_id] = [{
                    'move_id': ['N/A', 'N/A', 'N/A'],
                    'name': 'N/A',
                    'date': 'N/A',
                    'quantity': required_quantity,
                    'price_unit': 'N/A',
                    'currency_id': ['N/A', 'N/A']
                }]
    print()
    return result